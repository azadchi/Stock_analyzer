"""
اپلیکیشن تحلیل بورس
اجرا: python app.py  →  http://127.0.0.1:5000
نصب: pip install flask google-genai pillow openpyxl
تنظیم: $env:GEMINI_API_KEY="your_key"
"""

import os
import io
import base64
from datetime import datetime
from flask import Flask, request, render_template_string, send_file, redirect
from google import genai
from PIL import Image
import openpyxl

app = Flask(__name__)
HISTORY_FILE = "analysis_history.xlsx"
MODEL = "gemini-2.0-flash"

ANALYSIS_PROMPT = """
تو نقش یک مدیر سرمایه‌گذاری حرفه‌ای صندوق بورس ایران رو داری.

لطفاً:
1. تمام اعداد و اندیکاتورهای قابل مشاهده رو دقیق استخراج کن (قیمت فعلی، Ichimoku، RSI، MACD، MFI، OBV، ATR، حجم)
2. وضعیت روند، حمایت و مقاومت رو مشخص کن
3. برای هر اندیکاتور یک تحلیل کوتاه بده
4. نقطه ورود پیشنهادی، حد ضرر (بر اساس ATR)، و اهداف قیمتی رو بگو
5. جمع‌بندی نهایی: خرید | صبر و بررسی بیشتر | فروش

توجه: این فقط جنبه آموزشی داره و توصیه مالی نیست.
"""

BUY_PROMPT = """
تو یک تحلیلگر تکنیکال حرفه‌ای بورس ایران هستی. فقط روی نقطه خرید تمرکز کن:

1. بهترین قیمت ورود (دقیق)
2. حد ضرر (با دلیل)
3. هدف اول، دوم و سوم قیمتی
4. درصد احتمال موفقیت معامله
5. توصیه نهایی: همین الان بخر | صبر کن | نخر

توجه: این فقط جنبه آموزشی داره و توصیه مالی نیست.
"""

PAGE = """
<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
<meta charset="UTF-8">
<title>تحلیل بورس</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: Tahoma, Arial, sans-serif;
    background: linear-gradient(135deg, #0f0c29, #302b63, #24243e);
    min-height: 100vh; padding: 2rem; direction: rtl; color: #e0e0e0;
  }
  .container {
    max-width: 800px; margin: 0 auto;
    background: rgba(255,255,255,0.05);
    border-radius: 16px; padding: 2rem;
    backdrop-filter: blur(10px);
    border: 1px solid rgba(255,255,255,0.1);
  }
  h1 { font-size: 24px; margin-bottom: 0.25rem; color: #fff; }
  p.sub { color: #aaa; margin-bottom: 1.5rem; font-size: 14px; }
  .fields { display: grid; grid-template-columns: 1fr 1fr; gap: 1rem; margin-bottom: 1rem; }
  .fields input, .fields select, textarea {
    background: rgba(255,255,255,0.08); border: 1px solid rgba(255,255,255,0.2);
    border-radius: 8px; padding: 0.6rem 0.8rem; color: #fff; font-size: 14px; width: 100%;
  }
  textarea { width: 100%; margin-bottom: 1rem; resize: vertical; }
  .dropzone {
    border: 2px dashed rgba(255,255,255,0.3); border-radius: 10px;
    padding: 1.5rem; text-align: center; color: #aaa; margin-bottom: 1rem;
  }
  .preview-grid { display: flex; flex-wrap: wrap; gap: 0.5rem; margin-bottom: 1rem; }
  .preview-grid img { height: 80px; border-radius: 6px; border: 1px solid rgba(255,255,255,0.2); }
  .btns { display: flex; gap: 1rem; flex-wrap: wrap; }
  button {
    padding: 0.7rem 1.5rem; border: none; border-radius: 8px;
    font-size: 15px; cursor: pointer; font-family: Tahoma;
  }
  .btn-analyze { background: #2563eb; color: white; }
  .btn-analyze:hover { background: #1d4ed8; }
  .btn-buy { background: #16a34a; color: white; }
  .btn-buy:hover { background: #15803d; }
  .btn-history { background: rgba(255,255,255,0.1); color: #ddd; border: 1px solid rgba(255,255,255,0.2); }
  .result {
    margin-top: 1.5rem; background: rgba(0,0,0,0.3);
    border-radius: 10px; padding: 1.25rem;
    white-space: pre-wrap; line-height: 1.9; font-size: 14.5px;
    border: 1px solid rgba(255,255,255,0.1);
  }
  .error { color: #f87171; background: rgba(239,68,68,0.1); padding: 1rem; border-radius: 8px; margin-top: 1rem; }
  .disclaimer { font-size: 12px; color: #666; margin-top: 1.5rem; text-align: center; }
  a { color: #60a5fa; }
</style>
</head>
<body>
<div class="container">
  <h1>📊 تحلیل بورس</h1>
  <p class="sub">چارت تکنیکال سهم رو آپلود کن تا تحلیل کامل بگیری</p>

  <form method="POST" enctype="multipart/form-data">
    <div class="fields">
      <input type="text" name="symbol" placeholder="نماد سهم (مثلاً: شبندر)" value="{{ symbol or '' }}">
      <input type="text" name="timeframe" placeholder="تایم‌فریم (مثلاً: روزانه)" value="{{ timeframe or '' }}">
    </div>
    <textarea name="notes" rows="2" placeholder="یادداشت‌های جانبی (اختیاری)">{{ notes or '' }}</textarea>
    <div class="dropzone">
      تصاویر چارت رو انتخاب کن (می‌تونی چند تا انتخاب کنی)
      <br><input type="file" name="chart_images" accept="image/*" multiple>
    </div>
    {% if previews %}
    <div class="preview-grid">
      {% for p in previews %}
      <img src="data:image/png;base64,{{ p }}">
      {% endfor %}
    </div>
    {% endif %}
    <div class="btns">
      <button type="submit" name="action" value="analyze" class="btn-analyze">📈 تحلیل کامل</button>
      <button type="submit" name="action" value="buy" class="btn-buy">🎯 بهترین موقعیت خرید</button>
      <a href="/history"><button type="button" class="btn-history">📋 تاریخچه</button></a>
    </div>
  </form>

  {% if error %}
  <div class="error">خطا: {{ error }}</div>
  {% endif %}

  {% if result %}
  <div class="result">{{ result }}</div>
  {% endif %}

  <div class="disclaimer">این تحلیل جنبه آموزشی دارد و توصیه مالی قطعی نیست.</div>
</div>
</body>
</html>
"""

HISTORY_PAGE = """
<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
<meta charset="UTF-8"><title>تاریخچه تحلیل‌ها</title>
<style>
  body { font-family: Tahoma; background: linear-gradient(135deg,#0f0c29,#302b63,#24243e); min-height:100vh; padding:2rem; direction:rtl; color:#e0e0e0; }
  .container { max-width:900px; margin:0 auto; background:rgba(255,255,255,0.05); border-radius:16px; padding:2rem; border:1px solid rgba(255,255,255,0.1); }
  h1 { color:#fff; margin-bottom:1rem; }
  table { width:100%; border-collapse:collapse; font-size:13px; }
  th { background:rgba(255,255,255,0.1); padding:0.6rem; text-align:right; }
  td { padding:0.6rem; border-bottom:1px solid rgba(255,255,255,0.05); vertical-align:top; }
  .result-cell { max-width:400px; white-space:pre-wrap; font-size:12px; color:#ccc; }
  a { color:#60a5fa; }
  .btns { margin-bottom:1rem; display:flex; gap:1rem; }
  button { padding:0.5rem 1rem; border:none; border-radius:6px; cursor:pointer; font-family:Tahoma; }
  .btn-back { background:rgba(255,255,255,0.1); color:#ddd; border:1px solid rgba(255,255,255,0.2); }
  .btn-dl { background:#2563eb; color:white; }
</style>
</head>
<body>
<div class="container">
  <h1>📋 تاریخچه تحلیل‌ها</h1>
  <div class="btns">
    <a href="/"><button class="btn-back">← برگشت</button></a>
    <a href="/download"><button class="btn-dl">⬇ دانلود اکسل</button></a>
  </div>
  {% if rows %}
  <table>
    <tr><th>تاریخ</th><th>نماد</th><th>تایم‌فریم</th><th>نوع</th><th>نتیجه</th></tr>
    {% for r in rows %}
    <tr>
      <td>{{ r[0] }}</td><td>{{ r[1] }}</td><td>{{ r[2] }}</td><td>{{ r[3] }}</td>
      <td class="result-cell">{{ r[4][:300] }}{% if r[4]|length > 300 %}...{% endif %}</td>
    </tr>
    {% endfor %}
  </table>
  {% else %}
  <p style="color:#aaa">هنوز تحلیلی ثبت نشده.</p>
  {% endif %}
</div>
</body>
</html>
"""


def save_to_history(symbol, timeframe, analysis_type, result):
    if os.path.exists(HISTORY_FILE):
        wb = openpyxl.load_workbook(HISTORY_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["تاریخ", "نماد", "تایم‌فریم", "نوع تحلیل", "نتیجه"])
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M"), symbol, timeframe, analysis_type, result])
    wb.save(HISTORY_FILE)


@app.route("/", methods=["GET", "POST"])
def index():
    result = error = None
    previews = []
    symbol = timeframe = notes = ""

    if request.method == "POST":
        symbol = request.form.get("symbol", "")
        timeframe = request.form.get("timeframe", "")
        notes = request.form.get("notes", "")
        action = request.form.get("action", "analyze")

        if not os.environ.get("GEMINI_API_KEY"):
            error = "متغیر GEMINI_API_KEY تنظیم نشده."
        else:
            files = request.files.getlist("chart_images")
            files = [f for f in files if f and f.filename]
            if not files:
                error = "هیچ فایلی انتخاب نشد."
            else:
                try:
                    images = []
                    for f in files:
                        img = Image.open(f.stream)
                        images.append(img)
                        buf = io.BytesIO()
                        img.save(buf, format="PNG")
                        previews.append(base64.b64encode(buf.getvalue()).decode())

                    prompt = BUY_PROMPT if action == "buy" else ANALYSIS_PROMPT
                    if symbol:
                        prompt = f"نماد: {symbol}\n" + prompt
                    if timeframe:
                        prompt = f"تایم‌فریم: {timeframe}\n" + prompt
                    if notes:
                        prompt += f"\n\nیادداشت‌های اضافی: {notes}"

                    client = genai.Client()
                    contents = [prompt] + images
                    response = client.models.generate_content(model=MODEL, contents=contents)
                    result = response.text

                    analysis_type = "موقعیت خرید" if action == "buy" else "تحلیل کامل"
                    save_to_history(symbol, timeframe, analysis_type, result)
                except Exception as e:
                    error = str(e)

    return render_template_string(PAGE, result=result, error=error,
                                  previews=previews, symbol=symbol,
                                  timeframe=timeframe, notes=notes)


@app.route("/history")
def history():
    rows = []
    if os.path.exists(HISTORY_FILE):
        wb = openpyxl.load_workbook(HISTORY_FILE)
        ws = wb.active
        for row in list(ws.iter_rows(values_only=True))[1:]:
            rows.append(row)
    rows.reverse()
    return render_template_string(HISTORY_PAGE, rows=rows)


@app.route("/download")
def download():
    if os.path.exists(HISTORY_FILE):
        return send_file(HISTORY_FILE, as_attachment=True)
    return redirect("/history")


if __name__ == "__main__":
    print("سرور روشن شد → http://127.0.0.1:5000")
    app.run(debug=True, port=5000)
